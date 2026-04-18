"""
Store Review — data extraction, AI generation, and Excel output.
"""

from __future__ import annotations

import io
import json
import warnings
from datetime import datetime, timedelta
from pathlib import Path

warnings.filterwarnings("ignore")

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

APP_DIR   = Path(__file__).resolve().parent
DATA_DIR  = APP_DIR.parent / "data" / "reports"
VAULT_ROOT = APP_DIR.parent.parent.parent   # c:\Users\jason\vscode-vault
TEMPLATE  = (
    VAULT_ROOT
    / ".claude/skills/document-recreator/references"
    / "STORE REVIEW FORM Lebanon 3.24.xlsx"
)
VAULT_OUTPUTS = VAULT_ROOT / "outputs"

# Latest-report filenames (saved when user uploads)
SALES_FILE    = DATA_DIR / "retail-metrics-latest.xlsx"
RACK_FILE     = DATA_DIR / "rack-units-latest.xlsx"
TURNOVER_FILE = DATA_DIR / "turnover-latest.xlsx"

# ---------------------------------------------------------------------------
# Store roster
# ---------------------------------------------------------------------------

STORE_ROSTER = {
    "Tri-County":   {"sm": "Randy Conyers",      "asm": ["William Harper", "Michelle Baird"], "rack_key": "WOODLAWN"},
    "Lebanon":      {"sm": "Nancy Medina",        "asm": ["Lindsey Post"],                    "rack_key": "LEBANON"},
    "Loveland":     {"sm": "Amanda Leece",        "asm": ["Brandi Carman"],                   "rack_key": "LOVELAND"},
    "Fairfield":    {"sm": "Kimra Cappelletty",   "asm": ["Kristin Lewis", "Nadeisha Cummings"], "rack_key": "FAIRFIELD"},
    "Beechmont":    {"sm": "Donna Clouse",        "asm": ["Kathy Blanton"],                   "rack_key": "BEECHMONT"},
    "Deerfield":    {"sm": "Natiliia Fanfurniuk", "asm": ["Michael Mersman"],                 "rack_key": "DEERFIELD"},
    "ShopGoodwill": {"sm": "Vanessa Sevier",      "asm": ["Christian Jones"],                 "rack_key": None},
}

TURNOVER_NAME_MAP = {
    "Tri-County":   "tri-county store",
    "Lebanon":      "lebanon store",
    "Loveland":     "loveland store",
    "Fairfield":    "fairfield store",
    "Beechmont":    "beechmont store",
    "Deerfield":    "deerfield store",
    "ShopGoodwill": "shop goodwill",
}

# ---------------------------------------------------------------------------
# Metric extraction
# ---------------------------------------------------------------------------

def extract_sales_metrics(path: Path, store: str) -> dict:
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col = {str(h): i for i, h in enumerate(headers) if h}

    store_lower = store.strip().lower()
    for row in ws.iter_rows(min_row=2, values_only=True):
        loc = str(row[col.get("Location Name", 1)] or "").strip().lower()
        if store_lower in loc or loc in store_lower:
            return {
                "mtd_sales":          row[col["MTD Sales"]],
                "mtd_budget":         row[col["MTD Budget"]],
                "rev_pct_to_budget":  row[col["% MTD"]],
                "rev_pct_to_ly":      row[col["% MTD LY"]],
                "avg_transaction":    row[col["$/Trans"]],
                "mtd_ly":             row[col["MTD LY"]],
            }
    return {}


def extract_rack_metrics(path: Path, store: str) -> dict:
    rack_key = STORE_ROSTER.get(store, {}).get("rack_key") or store.upper()
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb["Rack Units Report"]

    for row in ws.iter_rows(min_row=7, max_row=28, values_only=True):
        cell = str(row[0] or "").strip().upper()
        if rack_key in cell or cell in rack_key:
            unit_goal        = row[4]
            units_on_hand    = row[5]
            units_pulled     = row[6]
            units_sold       = row[7]
            units_over_under = row[12]
            sell_thru        = row[19]

            if units_over_under is None and unit_goal and units_on_hand:
                calc = (
                    (units_on_hand or 0)
                    + (row[8] or 0) + (row[9] or 0) + (row[10] or 0)
                    - (units_pulled or 0) - (units_sold or 0)
                )
                units_over_under = calc - unit_goal

            return {
                "unit_goal":        unit_goal,
                "units_on_hand":    units_on_hand,
                "units_sold":       units_sold,
                "units_over_under": units_over_under,
                "sell_thru_pct":    sell_thru,
            }
    return {}


def extract_turnover_metrics(path: Path, store: str) -> dict:
    key = TURNOVER_NAME_MAP.get(store, store.lower() + " store")
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        dept = str(row[0] or "").strip().lower()
        if key in dept or dept in key:
            return {
                "ytd_turnover":        row[6],
                "current_employees":   row[2],
                "current_month_terms": row[9],
            }
    return {}


def load_all_metrics(store: str) -> dict:
    metrics = {}
    if SALES_FILE.exists():
        metrics.update(extract_sales_metrics(SALES_FILE, store))
    if RACK_FILE.exists():
        metrics.update(extract_rack_metrics(RACK_FILE, store))
    if TURNOVER_FILE.exists():
        metrics.update(extract_turnover_metrics(TURNOVER_FILE, store))
    return metrics


def save_report(uploaded_bytes: bytes, report_type: str):
    """Save uploaded report bytes to data/reports/."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    targets = {
        "sales":    SALES_FILE,
        "rack":     RACK_FILE,
        "turnover": TURNOVER_FILE,
    }
    targets[report_type].write_bytes(uploaded_bytes)


def reports_status() -> dict:
    """Return which reports are on file and their modification dates."""
    def info(p: Path):
        if p.exists():
            mtime = datetime.fromtimestamp(p.stat().st_mtime)
            return mtime.strftime("%m/%d/%Y %I:%M %p")
        return None
    return {
        "sales":    info(SALES_FILE),
        "rack":     info(RACK_FILE),
        "turnover": info(TURNOVER_FILE),
    }

# ---------------------------------------------------------------------------
# AI generation
# ---------------------------------------------------------------------------

def generate_store_review(client, store: str, metrics: dict, manual: dict) -> dict:
    """Call Claude to generate store comments and 3 priorities."""
    roster = STORE_ROSTER.get(store, {})
    sm   = roster.get("sm", "the store manager")
    asms = roster.get("asm", [])
    asm_str = " and ".join(asms) if asms else "the assistant manager"

    # Format key metrics for the prompt
    def pct(v):
        if v is None:
            return "N/A"
        return f"{'+' if v >= 0 else ''}{v*100:.1f}%"

    def num(v, decimals=0):
        if v is None:
            return "N/A"
        return f"{v:,.{decimals}f}"

    mtd_sales      = num(metrics.get("mtd_sales"), 0)
    mtd_budget     = num(metrics.get("mtd_budget"), 0)
    rev_budget     = pct(metrics.get("rev_pct_to_budget"))
    rev_ly         = pct(metrics.get("rev_pct_to_ly"))
    sell_thru      = pct(metrics.get("sell_thru_pct"))
    units_deficit  = metrics.get("units_over_under")
    unit_goal      = metrics.get("unit_goal")
    turnover       = pct(metrics.get("ytd_turnover"))
    avg_trans      = manual.get("avg_transaction_vs_ly") or metrics.get("avg_transaction")
    labor_raw      = manual.get("labor_pct_to_budget")
    labor          = pct(labor_raw)
    labor_status   = (
        "under 30% — on target" if labor_raw is not None and labor_raw < 0.30
        else "between 30-32% — watch" if labor_raw is not None and labor_raw <= 0.32
        else "above 32% — concern" if labor_raw is not None
        else "N/A"
    )
    donation       = pct(manual.get("donation_growth_pct_ly"))
    open_pos       = manual.get("open_positions", 0)
    ecomm          = pct(manual.get("ecommerce_pct_to_budget"))
    totes_mtd      = manual.get("ecommerce_totes_mtd", "N/A")
    carry_overs    = manual.get("carry_over_priorities", "")
    visit_date     = manual.get("visit_date", datetime.now().strftime("%m/%d/%Y"))
    notes          = manual.get("additional_notes", "")

    # Build deadline ~2-3 weeks out
    try:
        visit_dt = datetime.strptime(visit_date, "%m/%d/%Y")
        deadline = (visit_dt + timedelta(weeks=2)).strftime("%m/%d/%y").lstrip("0")
    except Exception:
        deadline = "EOM"

    # Yes/No observation summary
    obs_lines = []
    merch_map = {
        "merch_wares_full": "Wares Shelves Full",
        "merch_racks_full": "Racks Full",
        "merch_shoe_racks_full": "Shoe Racks Full",
        "merch_end_caps_full": "End Caps Full",
    }
    clean_map = {
        "clean_parking_lot": "Parking Lot",
        "clean_donation_area": "Donation Area",
        "clean_windows_entry": "Windows/Entry",
        "clean_cash_wraps": "Cash Wraps",
        "clean_floor_racks": "Floor/Racks",
        "clean_wares_shelves": "Wares/Shelves",
        "clean_fitting_rooms": "Fitting Rooms",
        "clean_production_room": "Production Room",
        "clean_offices": "Offices",
    }
    pull_map = {
        "pull_womens": "Women's", "pull_mens": "Men's", "pull_kids": "Kids",
        "pull_wares": "Wares", "pull_shoes": "Shoes", "pull_books": "Books", "pull_em": "E&M",
    }

    def obs_summary(mapping, label):
        nos = [v for k, v in mapping.items() if manual.get(k, "").lower() == "no"]
        if nos:
            obs_lines.append(f"{label} gaps: {', '.join(nos)}")

    obs_summary(merch_map,  "Merchandising")
    obs_summary(clean_map,  "Cleanliness")
    obs_summary(pull_map,   "Pull/Rotation")

    obs_text = "\n".join(obs_lines) if obs_lines else "No observation gaps noted."

    prompt = "\n".join([
        f"Write a store review for {store}. Use ONLY the data below. Do NOT mention any other store by name.",
        "",
        f"Store: {store}",
        f"Store Manager: {sm}",
        f"Assistant Manager(s): {asm_str}",
        f"Visit Date: {visit_date}",
        "",
        "--- STORE DATA ---",
        f"MTD Sales: ${mtd_sales} vs Budget ${mtd_budget} ({rev_budget} to budget, {rev_ly} to LY)",
        f"30-Day Sell Through: {sell_thru}",
        f"Unit Goal Deficit: {num(units_deficit, 0)} units below goal (goal is {num(unit_goal, 0)} units)",
        f"Average Transaction vs LY: ${avg_trans if avg_trans else 'N/A'}",
        f"Labor % to Budget: {labor} ({labor_status}) [goal: under 30%, concern: above 32%]",
        f"Donation Growth % to LY: {donation}",
        f"YTD Annualized Turnover: {turnover}",
        f"Open Positions: {open_pos}",
        f"Ecommerce % to Budget: {ecomm}",
        f"Ecommerce Totes MTD: {totes_mtd}",
        "",
        "Observation gaps (scored No during visit):",
        obs_text,
        "",
        f"Carry-over items from last visit: {carry_overs if carry_overs else 'None.'}",
        f"Additional notes: {notes if notes else 'None.'}",
        "",
        "--- INSTRUCTIONS ---",
        "",
        "GENERAL STORE COMMENTS (300-400 words):",
        f"- First sentence: MTD sales vs budget and vs LY with actual dollar amounts for {store}",
        f"- Always use the store name '{store}' and the manager first names {sm} and {asm_str} by first name",
        "- Keep it simple and direct - short sentences, plain language",
        "- CALL OUT POSITIVES EXPLICITLY: if sales are above budget or above LY, say so clearly and give credit to the team",
        "- If sell-through is strong, call it out as a win",
        "- If labor is under 30%, call it out as a positive — labor is under control",
        "- If ecommerce is above budget, highlight it",
        "- If donation growth is positive, mention it",
        "- If all cleanliness/merch/pull items are Yes, acknowledge the store looks sharp",
        "- THEN address what needs improvement: production output, unit goal deficit, labor if above 30%, staffing gaps",
        "- Labor context: under 30% = good, 30-32% = watch it, above 32% = flag as a concern",
        "- End with 4-5 bullet focus areas that preview the priorities",
        "- Do NOT mention any other store",
        "",
        "THREE PRIORITIES:",
        "- Focus on the biggest issues: production/unit deficit first, then staffing, then labor, then department gaps",
        f"- Opportunity: state the specific problem with a number and a deadline (~{deadline})",
        "- If carry-over: add 'Continued from previous visit.'",
        "- Action plan: 'Responsible: [name]' then 3-5 plain bullet steps",
        f"- Use {sm} and {asm_str} by name",
        "",
        "Return JSON only, no markdown:",
        "{",
        '  "comments": "...",',
        '  "priorities": [',
        f'    {{"opportunity": "...", "action_plan": "Responsible: {sm}\\n- step\\n- step"}},',
        '    {"opportunity": "...", "action_plan": "..."},',
        '    {"opportunity": "...", "action_plan": "..."}',
        "  ]",
        "}",
    ])

    message = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=2500,
        messages=[{"role": "user", "content": prompt}]
    )

    raw = message.content[0].text.strip()
    # Strip markdown code fences if present
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1]
        raw = raw.rsplit("```", 1)[0]

    return json.loads(raw)


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def _write_cell(ws, row: int, col: int, value):
    """Write value to cell, finding the merge master if needed."""
    cell = ws.cell(row=row, column=col)
    for mr in ws.merged_cells.ranges:
        if cell.coordinate in mr:
            ws.cell(row=mr.min_row, column=mr.min_col).value = value
            return
    cell.value = value


def build_store_review_xlsx(
    store: str,
    visit_date: str,
    prev_date: str,
    metrics: dict,
    manual: dict,
    ai_content: dict,
) -> bytes:
    wb = openpyxl.load_workbook(str(TEMPLATE))
    ws = wb.active

    visit_dt = datetime.strptime(visit_date, "%m/%d/%Y")
    prev_dt  = datetime.strptime(prev_date,  "%m/%d/%Y") if prev_date else None

    # ── Header ──────────────────────────────────────────────────────────────
    _write_cell(ws, 2, 2, store)
    _write_cell(ws, 4, 2, visit_dt)
    ws.cell(row=4, column=2).number_format = "M/D/YYYY"
    if prev_dt:
        _write_cell(ws, 5, 2, prev_dt)
        ws.cell(row=5, column=2).number_format = "M/D/YYYY"

    # ── Training ─────────────────────────────────────────────────────────────
    if manual.get("new_hires_met") is not None:
        _write_cell(ws, 4, 8, manual["new_hires_met"])
    if manual.get("num_new_hires") is not None:
        _write_cell(ws, 5, 8, manual["num_new_hires"])

    # ── General Metrics ───────────────────────────────────────────────────────
    # Write employees directly to L3 (bypass merge detection — L3 is inside K2:N3
    # which holds the "General Metrics" label at K2; _write_cell would overwrite it)
    ws.cell(row=3, column=12).value = manual.get("employees_spoken_to") or 0
    huddle = manual.get("huddles", "")
    if manual.get("huddles_days_missing"):
        huddle = f"{manual['huddles_days_missing']} days missing"
    _write_cell(ws, 4, 12, huddle or "Complete")
    if manual.get("ecommerce_pct_to_budget") is not None:
        _write_cell(ws, 7, 12, manual["ecommerce_pct_to_budget"])
    if manual.get("ecommerce_totes_mtd") is not None:
        _write_cell(ws, 8, 12, manual["ecommerce_totes_mtd"])
    if manual.get("truck_loaded"):
        _write_cell(ws, 9, 12, manual["truck_loaded"])
    if manual.get("rolling_racks_checked") is not None:
        _write_cell(ws, 3, 14, manual["rolling_racks_checked"])
    if manual.get("rolling_rack_items_back") is not None:
        _write_cell(ws, 4, 14, manual["rolling_rack_items_back"])
    if manual.get("hang_count_variance") is not None:
        _write_cell(ws, 5, 14, manual["hang_count_variance"])
    if manual.get("tote_count_variance") is not None:
        _write_cell(ws, 7, 14, manual["tote_count_variance"])
    if manual.get("totes_checked") is not None:
        _write_cell(ws, 8, 14, manual["totes_checked"])
    if manual.get("tote_items_back") is not None:
        _write_cell(ws, 9, 14, manual["tote_items_back"])

    # ── Pre-Visit Data ────────────────────────────────────────────────────────
    if metrics.get("rev_pct_to_budget") is not None:
        _write_cell(ws, 8, 3, metrics["rev_pct_to_budget"])
    if metrics.get("rev_pct_to_ly") is not None:
        _write_cell(ws, 9, 3, metrics["rev_pct_to_ly"])
    if manual.get("donation_growth_pct_ly") is not None:
        _write_cell(ws, 10, 3, manual["donation_growth_pct_ly"])
    if metrics.get("sell_thru_pct") is not None:
        _write_cell(ws, 11, 3, metrics["sell_thru_pct"])
    if manual.get("avg_transaction_vs_ly") is not None:
        _write_cell(ws, 8, 8, manual["avg_transaction_vs_ly"])
    if manual.get("labor_pct_to_budget") is not None:
        _write_cell(ws, 9, 8, manual["labor_pct_to_budget"])
    if metrics.get("ytd_turnover") is not None:
        _write_cell(ws, 10, 8, metrics["ytd_turnover"])
    if manual.get("open_positions") is not None:
        _write_cell(ws, 11, 8, manual["open_positions"])

    # ── Merchandising Yes/No (col D=4) ─────────────────────────────────────
    for row, key in [(14, "merch_wares_full"), (15, "merch_racks_full"),
                     (16, "merch_shoe_racks_full"), (17, "merch_end_caps_full")]:
        if manual.get(key):
            _write_cell(ws, row, 4, manual[key])

    # ── Cleanliness Yes/No (col D=4) ──────────────────────────────────────
    for row, key in [(20, "clean_parking_lot"), (21, "clean_donation_area"),
                     (22, "clean_windows_entry"), (23, "clean_cash_wraps"),
                     (24, "clean_floor_racks"), (25, "clean_wares_shelves"),
                     (26, "clean_fitting_rooms"), (27, "clean_production_room"),
                     (28, "clean_offices")]:
        if manual.get(key):
            _write_cell(ws, row, 4, manual[key])

    # ── Pull/Rotation Yes/No (col I=9) ─────────────────────────────────────
    # Labels are merged F:H (cols 6-8) — write Yes/No to col I (9)
    for row, key in [(15, "pull_womens"), (16, "pull_mens"), (17, "pull_kids"),
                     (18, "pull_wares"), (19, "pull_shoes"), (20, "pull_books"),
                     (21, "pull_em")]:
        if manual.get(key):
            ws.cell(row=row, column=9).value = manual[key]

    # ── Error Rate Racks (col M=13) ────────────────────────────────────────
    for row, key in [(14, "error_total_items"), (15, "error_wrong_price"),
                     (16, "error_defective"), (17, "error_wrong_category")]:
        if manual.get(key) is not None:
            _write_cell(ws, row, 13, manual[key])

    # ── Error Rate Containers (col M=13) ──────────────────────────────────
    for row, key in [(22, "container_total"), (23, "container_not_labeled"),
                     (24, "container_salvage_as_raw"), (25, "container_raw_as_salvage"),
                     (26, "container_soft_as_hard"), (27, "container_hard_as_soft")]:
        if manual.get(key) is not None:
            _write_cell(ws, row, 13, manual[key])

    # ── AI Content ────────────────────────────────────────────────────────
    if ai_content.get("comments"):
        _write_cell(ws, 31, 1, ai_content["comments"])  # Row 31 = merged A31:N37

    priority_rows = [(40, 42), (46, 48), (52, 54)]
    for i, (opp_row, plan_row) in enumerate(priority_rows):
        plist = ai_content.get("priorities", [])
        if i < len(plist):
            p = plist[i]
            _write_cell(ws, opp_row, 4, p.get("opportunity", ""))
            _write_cell(ws, plan_row, 4, p.get("action_plan", ""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
