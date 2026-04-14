"""
Extract store metrics from uploaded data reports.

Usage:
  python extract_metrics.py --store "Lebanon" \
    --sales "path/to/Retail Metrics by District.xlsx" \
    [--rack_units "path/to/Rack Units Report.xlsx"] \
    [--turnover "path/to/Employee Turnover.xlsx"] \
    [--output "metrics.json"]
"""

import argparse
import json
import warnings
warnings.filterwarnings("ignore")

import openpyxl

# Store name normalization maps
SALES_NAME_MAP = {
    "tri-county": "tri-county",
    "tricounty": "tri-county",
    "woodlawn": "tri-county",
    "lebanon": "lebanon",
    "loveland": "loveland",
    "fairfield": "fairfield",
    "beechmont": "beechmont",
    "deerfield": "deerfield",
    "shopgoodwill": "shopgoodwill",
    "shop goodwill": "shopgoodwill",
}

RACK_NAME_MAP = {
    "tri-county": "woodlawn",
    "woodlawn": "woodlawn",
    "lebanon": "lebanon",
    "loveland": "loveland",
    "fairfield": "fairfield",
    "beechmont": "beechmont",
    "deerfield": "deerfield",
}

TURNOVER_NAME_MAP = {
    "tri-county": "tri-county store",
    "woodlawn": "tri-county store",
    "lebanon": "lebanon store",
    "loveland": "loveland store",
    "fairfield": "fairfield store",
    "beechmont": "beechmont store",
    "deerfield": "deerfield store",
    "shopgoodwill": "shop goodwill",
}


def normalize(name):
    return name.strip().lower()


def extract_sales(path, store_key):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    col = {h: i for i, h in enumerate(headers) if h}

    for row in ws.iter_rows(min_row=2, values_only=True):
        loc_name = str(row[col.get("Location Name", 1)] or "").strip().lower()
        if store_key in loc_name or loc_name in store_key:
            return {
                "mtd_sales": row[col["MTD Sales"]],
                "mtd_budget": row[col["MTD Budget"]],
                "rev_pct_to_budget": row[col["% MTD"]],
                "rev_pct_to_ly": row[col["% MTD LY"]],
                "avg_transaction_dollar": row[col["$/Trans"]],
                "same_day_ly_transaction": None,  # calc below
            }
    return {}


def extract_rack_units(path, store_key):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Rack Units Report"]

    rack_key = RACK_NAME_MAP.get(store_key, store_key).upper().strip()

    for row in ws.iter_rows(min_row=7, max_row=28, values_only=True):
        store_cell = str(row[0] or "").strip().upper()
        if rack_key in store_cell or store_cell in rack_key:
            # Columns (0-indexed): A=store, B=pull, C=units/lf, D=linear_ft,
            # E=unit_goal(formula), F=units_on_hand, G=units_pulled,
            # H=units_sold, I=shipped, J=cons_prod, K=calc_prod,
            # L=calc_on_hand(formula), M=units_over_under(formula),
            # P=$_sold, T=sell_thru_2026, U=sell_thru_2025
            unit_goal = row[4]   # col E
            units_on_hand = row[5]   # col F
            units_pulled = row[6]   # col G
            units_sold = row[7]   # col H
            units_over_under = row[12]  # col M
            sell_thru = row[19]  # col T

            # units_over_under may be a formula result — compute manually if None
            if units_over_under is None and unit_goal and units_on_hand:
                calc_on_hand = (units_on_hand or 0) + (row[8] or 0) + (row[9] or 0) + (row[10] or 0) - (units_pulled or 0) - (units_sold or 0)
                units_over_under = calc_on_hand - unit_goal

            return {
                "unit_goal": unit_goal,
                "units_on_hand": units_on_hand,
                "units_pulled": units_pulled,
                "units_sold": units_sold,
                "units_over_under": units_over_under,
                "sell_thru_pct": sell_thru,
            }
    return {}


def extract_turnover(path, store_key):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    turnover_key = TURNOVER_NAME_MAP.get(store_key, store_key + " store")

    for row in ws.iter_rows(min_row=2, values_only=True):
        dept = str(row[0] or "").strip().lower()
        if turnover_key in dept or dept in turnover_key:
            return {
                "ytd_annualized_turnover": row[6],  # col G
                "current_employees": row[2],         # col C
                "current_month_terms": row[9],       # col J
            }
    return {}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--store", required=True)
    parser.add_argument("--sales", required=False)
    parser.add_argument("--rack_units", required=False)
    parser.add_argument("--turnover", required=False)
    parser.add_argument("--output", default="metrics.json")
    args = parser.parse_args()

    store_key = normalize(args.store)
    # Normalize to canonical key
    store_key = SALES_NAME_MAP.get(store_key, store_key)

    metrics = {"store": args.store}

    if args.sales:
        sales = extract_sales(args.sales, store_key)
        metrics.update(sales)
        if not sales:
            print(f"WARNING: Store '{args.store}' not found in sales report.")

    if args.rack_units:
        rack = extract_rack_units(args.rack_units, store_key)
        metrics.update(rack)
        if not rack:
            print(f"WARNING: Store '{args.store}' not found in rack units report.")

    if args.turnover:
        turnover = extract_turnover(args.turnover, store_key)
        metrics.update(turnover)
        if not turnover:
            print(f"WARNING: Store '{args.store}' not found in turnover report.")

    # Format percentages for display
    for key in ["rev_pct_to_budget", "rev_pct_to_ly", "sell_thru_pct", "ytd_annualized_turnover"]:
        if key in metrics and metrics[key] is not None:
            metrics[f"{key}_display"] = f"{metrics[key]*100:.1f}%"

    with open(args.output, "w") as f:
        json.dump(metrics, f, indent=2, default=str)

    print(f"Metrics extracted for {args.store}:")
    for k, v in metrics.items():
        if not k.endswith("_display"):
            print(f"  {k}: {v}")

    print(f"\nSaved to: {args.output}")


if __name__ == "__main__":
    main()
