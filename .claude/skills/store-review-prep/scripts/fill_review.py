"""
Fill the store review Excel template with extracted metrics, manual inputs,
and AI-generated comments/priorities.

Usage:
  python fill_review.py \
    --store "Lebanon" \
    --visit_date "2026-04-13" \
    --prev_date "2026-03-24" \
    --metrics "metrics.json" \
    --manual "manual.json" \
    --comments "comments.txt" \
    --priorities "priorities.json" \
    --output "outputs/store-reviews/2026-04-13_Lebanon_store-review.xlsx"

Manual JSON structure:
{
  "labor_pct_to_budget": 0.27,
  "donation_growth_pct_ly": -0.03,
  "open_positions": 2,
  "ecommerce_pct_to_budget": 0.16,
  "ecommerce_totes_mtd": 70,
  "avg_transaction_vs_ly": 2.44,
  "truck_loaded_correctly": "Yes",
  "huddles": "Complete",
  "huddles_days_missing": 0,
  "rolling_racks_checked": 2,
  "rolling_rack_items_sent_back": 0,
  "hang_count_variance": -0.15,
  "tote_count_variance": -0.09,
  "totes_checked": 4,
  "tote_items_sent_back": 0,
  "employees_spoken_to": "Nancy, Lindsey, Jacob",
  "new_hires_met": 1,
  "num_new_hires": 1,
  "merch_wares_full": "Yes",
  "merch_racks_full": "Yes",
  "merch_shoe_racks_full": "No",
  "merch_end_caps_full": "Yes",
  "clean_parking_lot": "Yes",
  "clean_donation_area": "Yes",
  "clean_windows_entry": "Yes",
  "clean_cash_wraps": "Yes",
  "clean_floor_racks": "Yes",
  "clean_wares_shelves": "Yes",
  "clean_fitting_rooms": "Yes",
  "clean_production_room": "No",
  "clean_offices": "Yes",
  "pull_womens": "Yes",
  "pull_mens": "Yes",
  "pull_kids": "Yes",
  "pull_wares": "No",
  "pull_shoes": "Yes",
  "pull_books": "Yes",
  "pull_em": "Yes",
  "error_total_items_checked": 80,
  "error_wrong_price": 2,
  "error_defective": 1,
  "error_wrong_category": 3,
  "container_total_checked": 17,
  "container_not_labeled": 2,
  "container_salvage_as_raw": 0,
  "container_raw_as_salvage": 0,
  "container_soft_as_hard": 0,
  "container_hard_as_soft": 0
}

Priorities JSON structure:
[
  {
    "opportunity": "Opportunity statement text",
    "action_plan": "Full SMART action plan text"
  },
  ...
]
"""

import argparse
import json
import shutil
import warnings
from datetime import datetime
from pathlib import Path

warnings.filterwarnings("ignore")

import openpyxl

TEMPLATE = Path(__file__).resolve().parents[3] / \
    ".claude/skills/document-recreator/references/STORE REVIEW FORM Lebanon 3.24.xlsx"


def write_cell(ws, row, col, value):
    """Write to cell by row/col, handles merged cells by finding the master cell."""
    cell = ws.cell(row=row, column=col)
    # If this cell is part of a merged range, find the top-left master cell
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            master = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            master.value = value
            return
    cell.value = value


def col_letter_to_num(letter):
    """Convert column letter to number (A=1, B=2, ...)."""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result


def fill_review(store, visit_date, prev_date, metrics, manual, comments, priorities, output_path):
    # Load template
    wb = openpyxl.load_workbook(str(TEMPLATE))
    ws = wb.active

    visit_dt = datetime.strptime(visit_date, "%Y-%m-%d")
    prev_dt = datetime.strptime(prev_date, "%Y-%m-%d") if prev_date else None

    # ── Header ──────────────────────────────────────────────────────────────
    # Row 2: Store name (col B=2)
    write_cell(ws, 2, 2, store)
    # Row 4: Visit date (col B=2)
    write_cell(ws, 4, 2, visit_dt)
    ws.cell(row=4, column=2).number_format = "M/D/YYYY"
    # Row 5: Previous eval date (col B=2)
    if prev_dt:
        write_cell(ws, 5, 2, prev_dt)
        ws.cell(row=5, column=2).number_format = "M/D/YYYY"

    # ── Training (col H=8) ──────────────────────────────────────────────────
    if manual.get("new_hires_met") is not None:
        write_cell(ws, 4, 8, manual["new_hires_met"])
    if manual.get("num_new_hires") is not None:
        write_cell(ws, 5, 8, manual["num_new_hires"])

    # ── General Metrics ──────────────────────────────────────────────────────
    # Employees spoken to (col L=12, row 3)
    if manual.get("employees_spoken_to"):
        write_cell(ws, 3, 12, manual["employees_spoken_to"])
    # Huddles (col L=12, row 4)
    if manual.get("huddles"):
        huddle_val = manual["huddles"]
        if manual.get("huddles_days_missing") and manual["huddles_days_missing"] > 0:
            huddle_val = f"{manual['huddles_days_missing']} days missing"
        write_cell(ws, 4, 12, huddle_val)
    # Ecommerce % to budget (col L=12, row 7)
    if metrics.get("ecommerce_pct_to_budget") is not None or manual.get("ecommerce_pct_to_budget") is not None:
        val = manual.get("ecommerce_pct_to_budget") or metrics.get("ecommerce_pct_to_budget")
        write_cell(ws, 7, 12, val)
    # Ecommerce totes MTD (col L=12, row 8)
    if manual.get("ecommerce_totes_mtd") is not None:
        write_cell(ws, 8, 12, manual["ecommerce_totes_mtd"])
    # Truck loaded (col L=12, row 9)
    if manual.get("truck_loaded_correctly"):
        write_cell(ws, 9, 12, manual["truck_loaded_correctly"])
    # Rolling racks checked (col N=14, row 3)
    if manual.get("rolling_racks_checked") is not None:
        write_cell(ws, 3, 14, manual["rolling_racks_checked"])
    # Rolling rack items sent back (col N=14, row 4) — approximate row
    if manual.get("rolling_rack_items_sent_back") is not None:
        write_cell(ws, 4, 14, manual["rolling_rack_items_sent_back"])
    # Hang count variance (col N=14, row 5 area — row 6 in template)
    if manual.get("hang_count_variance") is not None:
        write_cell(ws, 5, 14, manual["hang_count_variance"])
    # Tote count variance (col N=14, row 7)
    if manual.get("tote_count_variance") is not None:
        write_cell(ws, 7, 14, manual["tote_count_variance"])
    # Totes checked (col N=14, row 9 area)
    if manual.get("totes_checked") is not None:
        write_cell(ws, 8, 14, manual["totes_checked"])
    # Tote items sent back (col N=14, row 9)
    if manual.get("tote_items_sent_back") is not None:
        write_cell(ws, 9, 14, manual["tote_items_sent_back"])

    # ── Pre-Visit Data ──────────────────────────────────────────────────────
    # Rev % to budget (col C=3, row 8)
    if metrics.get("rev_pct_to_budget") is not None:
        write_cell(ws, 8, 3, metrics["rev_pct_to_budget"])
    # Rev % to LY (col C=3, row 9)
    if metrics.get("rev_pct_to_ly") is not None:
        write_cell(ws, 9, 3, metrics["rev_pct_to_ly"])
    # Donation growth % to LY (col C=3, row 10)
    val = manual.get("donation_growth_pct_ly") or metrics.get("donation_growth_pct_ly")
    if val is not None:
        write_cell(ws, 10, 3, val)
    # 30 day sell through (col C=3, row 11)
    if metrics.get("sell_thru_pct") is not None:
        write_cell(ws, 11, 3, metrics["sell_thru_pct"])
    # Avg transaction +/- to LY (col H=8, row 8)
    val = manual.get("avg_transaction_vs_ly") or metrics.get("avg_transaction_vs_ly")
    if val is not None:
        write_cell(ws, 8, 8, val)
    # Labor % to budget (col H=8, row 9)
    if manual.get("labor_pct_to_budget") is not None:
        write_cell(ws, 9, 8, manual["labor_pct_to_budget"])
    # Annualized turnover (col H=8, row 10)
    if metrics.get("ytd_annualized_turnover") is not None:
        write_cell(ws, 10, 8, metrics["ytd_annualized_turnover"])
    # Open positions (col H=8, row 11)
    if manual.get("open_positions") is not None:
        write_cell(ws, 11, 8, manual["open_positions"])

    # ── Merchandising Yes/No (col D=4) ──────────────────────────────────────
    merch_fields = [
        (13, "merch_wares_full"),
        (14, "merch_racks_full"),
        (15, "merch_shoe_racks_full"),
        (16, "merch_end_caps_full"),
    ]
    for row, key in merch_fields:
        if manual.get(key):
            write_cell(ws, row, 4, manual[key])

    # ── Store Cleanliness Yes/No (col D=4) ──────────────────────────────────
    clean_fields = [
        (18, "clean_parking_lot"),
        (19, "clean_donation_area"),
        (20, "clean_windows_entry"),
        (21, "clean_cash_wraps"),
        (22, "clean_floor_racks"),
        (23, "clean_wares_shelves"),
        (24, "clean_fitting_rooms"),
        (25, "clean_production_room"),
        (26, "clean_offices"),
    ]
    for row, key in clean_fields:
        if manual.get(key):
            write_cell(ws, row, 4, manual[key])

    # ── Pull/Rotation Validation Yes/No (col H=8) ───────────────────────────
    pull_fields = [
        (14, "pull_womens"),
        (15, "pull_mens"),
        (16, "pull_kids"),
        (17, "pull_wares"),
        (18, "pull_shoes"),
        (19, "pull_books"),
        (20, "pull_em"),
    ]
    for row, key in pull_fields:
        if manual.get(key):
            write_cell(ws, row, 8, manual[key])

    # ── Error Rate Racks (col M=13, rows 14-17) ─────────────────────────────
    error_fields = [
        (14, "error_total_items_checked"),
        (15, "error_wrong_price"),
        (16, "error_defective"),
        (17, "error_wrong_category"),
    ]
    for row, key in error_fields:
        if manual.get(key) is not None:
            write_cell(ws, row, 13, manual[key])

    # ── Error Rate Containers (col M=13, rows 22-27) ─────────────────────────
    container_fields = [
        (21, "container_total_checked"),
        (22, "container_not_labeled"),
        (23, "container_salvage_as_raw"),
        (24, "container_raw_as_salvage"),
        (25, "container_soft_as_hard"),
        (26, "container_hard_as_soft"),
    ]
    for row, key in container_fields:
        if manual.get(key) is not None:
            write_cell(ws, row, 13, manual[key])

    # ── General Store Comments (row 30, col A=1) ────────────────────────────
    if comments:
        write_cell(ws, 30, 1, comments)

    # ── Three Priorities ─────────────────────────────────────────────────────
    # Priority rows: 1=row40/42, 2=row46/48, 3=row52/54
    priority_rows = [
        (40, 42),   # Priority 1: opportunity row, action plan row
        (46, 48),   # Priority 2
        (52, 54),   # Priority 3
    ]
    for i, (opp_row, plan_row) in enumerate(priority_rows):
        if i < len(priorities):
            p = priorities[i]
            if p.get("opportunity"):
                write_cell(ws, opp_row, 4, p["opportunity"])
            if p.get("action_plan"):
                write_cell(ws, plan_row, 4, p["action_plan"])

    # ── Save ─────────────────────────────────────────────────────────────────
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    print(f"Saved: {out}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--store", required=True)
    parser.add_argument("--visit_date", required=True)
    parser.add_argument("--prev_date", default="")
    parser.add_argument("--metrics", default=None)
    parser.add_argument("--manual", default=None)
    parser.add_argument("--comments", default=None)
    parser.add_argument("--priorities", default=None)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    metrics = {}
    if args.metrics and Path(args.metrics).exists():
        with open(args.metrics) as f:
            metrics = json.load(f)

    manual = {}
    if args.manual and Path(args.manual).exists():
        with open(args.manual) as f:
            manual = json.load(f)

    comments = ""
    if args.comments:
        p = Path(args.comments)
        if p.exists():
            comments = p.read_text(encoding="utf-8")
        else:
            comments = args.comments  # passed as inline string

    priorities = []
    if args.priorities and Path(args.priorities).exists():
        with open(args.priorities) as f:
            priorities = json.load(f)

    fill_review(
        store=args.store,
        visit_date=args.visit_date,
        prev_date=args.prev_date,
        metrics=metrics,
        manual=manual,
        comments=comments,
        priorities=priorities,
        output_path=args.output,
    )


if __name__ == "__main__":
    main()
